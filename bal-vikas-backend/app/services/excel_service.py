from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class ExcelService:
    """Service for exporting assessment data to Excel matching the official ECD dataset format"""
    
    @staticmethod
    def export_child_assessment(assessment_data: Dict[str, Any], output_path: str) -> str:
        """
        Export child assessment data matching the exact ECD sample dataset format
        Format: Worked_Example_30_Month_Arjun.xlsx structure
        """
        wb = Workbook()
        
        # Sheet 1: A_Registration
        ws1 = wb.active
        ws1.title = "A_Registration"
        ExcelService._write_registration_sheet(ws1, assessment_data)
        
        # Sheet 2: Developmental_Assessment
        ws2 = wb.create_sheet("Developmental_Assessment")
        ExcelService._write_developmental_sheet(ws2, assessment_data)
        
        # Sheet 3: Developmental_Risk
        ws3 = wb.create_sheet("Developmental_Risk")
        ExcelService._write_risk_sheet(ws3, assessment_data)
        
        # Sheet 4: Neuro_Behavioral
        ws4 = wb.create_sheet("Neuro_Behavioral")
        ExcelService._write_neuro_sheet(ws4, assessment_data)
        
        # Sheet 5: Behaviour_Indicators
        ws5 = wb.create_sheet("Behaviour_Indicators")
        ExcelService._write_behavior_sheet(ws5, assessment_data)
        
        # Sheet 6: Environment_Caregiving
        ws6 = wb.create_sheet("Environment_Caregiving")
        ExcelService._write_environment_sheet(ws6, assessment_data)
        
        # Sheet 7: Nutrition
        ws7 = wb.create_sheet("Nutrition")
        ExcelService._write_nutrition_sheet(ws7, assessment_data)
        
        # Sheet 8: Baseline_Risk_Output
        ws8 = wb.create_sheet("Baseline_Risk_Output")
        ExcelService._write_baseline_sheet(ws8, assessment_data)
        
        wb.save(output_path)
        return output_path
    
    @staticmethod
    def _write_registration_sheet(ws, data: Dict[str, Any]):
        """Write A_Registration sheet - Basic child information"""
        headers = [
            "child_id", "child_unique_id", "age_months", "gender", 
            "anganwadi_center_id", "anganwadi_center", "district", "mandal", "village",
            "assessment_date", "assessor_name", "assessor_role"
        ]
        ws.append(headers)
        
        row = [
            data.get("child_id", ""),
            data.get("child_unique_id", ""),
            data.get("child_age_months", ""),
            data.get("gender", ""),
            data.get("anganwadi_center_id", ""),
            data.get("anganwadi_center", ""),
            data.get("district", ""),
            data.get("mandal", ""),
            data.get("village", ""),
            data.get("assessment_date", datetime.now().strftime("%Y-%m-%d")),
            data.get("assessor_name", ""),
            data.get("assessor_role", "")
        ]
        ws.append(row)
        ExcelService._format_sheet(ws)
    
    @staticmethod
    def _write_developmental_sheet(ws, data: Dict[str, Any]):
        """Write Developmental_Assessment sheet - DQ scores and birth history"""
        headers = [
            "child_id", "child_unique_id",
            "mode_delivery", "mode_conception", "birth_status", "consanguinity",
            "birth_weight_kg", "birth_complications",
            "GM_DQ", "FM_DQ", "LC_DQ", "COG_DQ", "SE_DQ", 
            "Composite_DQ",
            "GM_DA_months", "FM_DA_months", "LC_DA_months", "COG_DA_months", "SE_DA_months",
            "assessment_date"
        ]
        ws.append(headers)
        
        dev = data.get("developmental", {})
        age_months = data.get("child_age_months", 0)
        
        # Calculate developmental ages from DQ scores
        gm_dq = dev.get("gm_dq", 0) or 0
        fm_dq = dev.get("fm_dq", 0) or 0
        lc_dq = dev.get("lc_dq", 0) or 0
        cog_dq = dev.get("cog_dq", 0) or 0
        se_dq = dev.get("se_dq", 0) or 0
        
        gm_da = round((gm_dq * age_months) / 100, 1) if age_months > 0 else 0
        fm_da = round((fm_dq * age_months) / 100, 1) if age_months > 0 else 0
        lc_da = round((lc_dq * age_months) / 100, 1) if age_months > 0 else 0
        cog_da = round((cog_dq * age_months) / 100, 1) if age_months > 0 else 0
        se_da = round((se_dq * age_months) / 100, 1) if age_months > 0 else 0
        
        row = [
            data.get("child_id", ""),
            data.get("child_unique_id", ""),
            data.get("mode_delivery", ""),
            data.get("mode_conception", ""),
            data.get("birth_status", ""),
            data.get("consanguinity", ""),
            data.get("birth_weight_kg", ""),
            data.get("birth_complications", ""),
            gm_dq,
            fm_dq,
            lc_dq,
            cog_dq,
            se_dq,
            dev.get("composite_dq", ""),
            gm_da,
            fm_da,
            lc_da,
            cog_da,
            se_da,
            data.get("assessment_date", datetime.now().strftime("%Y-%m-%d"))
        ]
        ws.append(row)
        ExcelService._format_sheet(ws)
    
    @staticmethod
    def _write_risk_sheet(ws, data: Dict[str, Any]):
        """Write Developmental_Risk sheet - Delay flags"""
        headers = [
            "child_id", "child_unique_id",
            "GM_delay", "FM_delay", "LC_delay", "COG_delay", "SE_delay", 
            "num_delays",
            "delayed_domains",
            "delay_severity",  # None, Mild, Moderate, Severe
            "assessment_date"
        ]
        ws.append(headers)
        
        risk = data.get("risk", {})
        num_delays = risk.get("num_delays", 0)
        
        # Determine delayed domains
        delayed_domains = []
        if risk.get("gm_delay"):
            delayed_domains.append("Gross Motor")
        if risk.get("fm_delay"):
            delayed_domains.append("Fine Motor")
        if risk.get("lc_delay"):
            delayed_domains.append("Language & Communication")
        if risk.get("cog_delay"):
            delayed_domains.append("Cognitive")
        if risk.get("se_delay"):
            delayed_domains.append("Social-Emotional")
        
        # Determine delay severity
        if num_delays == 0:
            delay_severity = "None"
        elif num_delays == 1:
            delay_severity = "Mild"
        elif num_delays == 2:
            delay_severity = "Moderate"
        else:
            delay_severity = "Severe"
        
        row = [
            data.get("child_id", ""),
            data.get("child_unique_id", ""),
            "Yes" if risk.get("gm_delay") else "No",
            "Yes" if risk.get("fm_delay") else "No",
            "Yes" if risk.get("lc_delay") else "No",
            "Yes" if risk.get("cog_delay") else "No",
            "Yes" if risk.get("se_delay") else "No",
            num_delays,
            ", ".join(delayed_domains) if delayed_domains else "None",
            delay_severity,
            data.get("assessment_date", datetime.now().strftime("%Y-%m-%d"))
        ]
        ws.append(row)
        ExcelService._format_sheet(ws)
    
    @staticmethod
    def _write_neuro_sheet(ws, data: Dict[str, Any]):
        """Write Neuro_Behavioral sheet - Autism and behavioral screening"""
        headers = [
            "child_id", "child_unique_id",
            "autism_risk", "adhd_risk", "behavior_risk",
            "mchat_score", "mchat_critical_items_failed",
            "isaa_score", "adhd_score", 
            "sdq_total_score", "sdq_emotional", "sdq_conduct", 
            "sdq_hyperactivity", "sdq_peer", "sdq_prosocial",
            "assessment_date"
        ]
        ws.append(headers)
        
        neuro = data.get("neuro_behavioral", {})
        row = [
            data.get("child_id", ""),
            data.get("child_unique_id", ""),
            neuro.get("autism_risk", "Low"),
            neuro.get("adhd_risk", "Low"),
            neuro.get("behavior_risk", "Low"),
            neuro.get("mchat_score", ""),
            "",  # mchat_critical_items_failed - would need detailed data
            neuro.get("isaa_score", ""),
            neuro.get("adhd_score", ""),
            neuro.get("sdq_total_score", ""),
            "",  # sdq_emotional
            "",  # sdq_conduct
            "",  # sdq_hyperactivity
            "",  # sdq_peer
            "",  # sdq_prosocial
            data.get("assessment_date", datetime.now().strftime("%Y-%m-%d"))
        ]
        ws.append(row)
        ExcelService._format_sheet(ws)
    
    @staticmethod
    def _write_behavior_sheet(ws, data: Dict[str, Any]):
        """Write Behaviour_Indicators sheet - Detailed behavioral concerns"""
        headers = [
            "child_id", "child_unique_id",
            "behaviour_concerns_notes",
            "behaviour_score",
            "behaviour_risk_level",
            "emotional_symptoms",
            "conduct_problems",
            "hyperactivity",
            "peer_problems",
            "prosocial_behavior",
            "assessment_date"
        ]
        ws.append(headers)
        
        behavior = data.get("behavior", {})
        row = [
            data.get("child_id", ""),
            data.get("child_unique_id", ""),
            behavior.get("concerns", ""),
            behavior.get("score", ""),
            behavior.get("risk_level", "Low"),
            "",  # emotional_symptoms
            "",  # conduct_problems
            "",  # hyperactivity
            "",  # peer_problems
            "",  # prosocial_behavior
            data.get("assessment_date", datetime.now().strftime("%Y-%m-%d"))
        ]
        ws.append(row)
        ExcelService._format_sheet(ws)
    
    @staticmethod
    def _write_environment_sheet(ws, data: Dict[str, Any]):
        """Write Environment_Caregiving sheet - Home environment assessment"""
        headers = [
            "child_id", "child_unique_id",
            "parent_child_interaction_score",
            "home_stimulation_score",
            "play_materials_available",
            "caregiver_engagement",
            "language_exposure",
            "safe_water_available",
            "toilet_facility_available",
            "adequate_nutrition",
            "environment_total_score",
            "environment_max_score",
            "environment_percentage",
            "environment_risk_level",
            "assessment_date"
        ]
        ws.append(headers)
        
        env = data.get("environment", {})
        
        # Calculate environment percentage if scores available
        total = env.get("parent_child_interaction_score", 0) + env.get("home_stimulation_score", 0)
        max_score = 25
        percentage = round((total / max_score) * 100, 2) if max_score > 0 else 0
        
        row = [
            data.get("child_id", ""),
            data.get("child_unique_id", ""),
            env.get("parent_child_interaction_score", ""),
            env.get("home_stimulation_score", ""),
            "Yes" if env.get("play_materials") else "No",
            env.get("caregiver_engagement", ""),
            env.get("language_exposure", ""),
            "Yes" if env.get("safe_water") else "No",
            "Yes" if env.get("toilet_facility") else "No",
            "",  # adequate_nutrition
            total,
            max_score,
            percentage,
            "Low" if percentage >= 80 else "Medium" if percentage >= 60 else "High",
            data.get("assessment_date", datetime.now().strftime("%Y-%m-%d"))
        ]
        ws.append(row)
        ExcelService._format_sheet(ws)
    
    @staticmethod
    def _write_nutrition_sheet(ws, data: Dict[str, Any]):
        """Write Nutrition sheet - Anthropometric measurements"""
        headers = [
            "child_id", "child_unique_id",
            "height_cm", "weight_kg", "head_circumference_cm",
            "height_z_score", "weight_z_score", "wfh_z_score",
            "underweight_status",  # Normal, Moderate, Severe
            "stunting_status",     # Normal, Moderate, Severe
            "wasting_status",      # Normal, Moderate, Severe
            "anemia_status",       # 0=No, 1=Mild, 2=Severe
            "nutrition_score",
            "nutrition_risk",      # Low, Medium, High
            "who_chart_reference",
            "assessment_date"
        ]
        ws.append(headers)
        
        nutrition = data.get("nutrition", {})
        
        # Map numeric classifications to text
        underweight_map = {0: "Normal", 1: "Moderate", 2: "Severe"}
        stunting_map = {0: "Normal", 1: "Moderate", 2: "Severe"}
        wasting_map = {0: "Normal", 1: "Moderate", 2: "Severe"}
        anemia_map = {0: "No", 1: "Mild", 2: "Severe"}
        
        row = [
            data.get("child_id", ""),
            data.get("child_unique_id", ""),
            nutrition.get("height_cm", ""),
            nutrition.get("weight_kg", ""),
            nutrition.get("head_circumference_cm", ""),
            nutrition.get("height_z_score", ""),
            nutrition.get("weight_z_score", ""),
            nutrition.get("wfh_z_score", ""),
            underweight_map.get(nutrition.get("underweight", 0), "Normal"),
            stunting_map.get(nutrition.get("stunting", 0), "Normal"),
            wasting_map.get(nutrition.get("wasting", 0), "Normal"),
            anemia_map.get(nutrition.get("anemia", 0), "No"),
            nutrition.get("nutrition_score", ""),
            nutrition.get("nutrition_risk", ""),
            "WHO 2006 Child Growth Standards",
            data.get("assessment_date", datetime.now().strftime("%Y-%m-%d"))
        ]
        ws.append(row)
        ExcelService._format_sheet(ws)
    
    @staticmethod
    def _write_baseline_sheet(ws, data: Dict[str, Any]):
        """Write Baseline_Risk_Output sheet - Final risk classification"""
        headers = [
            "child_id", "child_unique_id",
            "overall_risk_category",  # LOW, MEDIUM, MEDIUM-HIGH, HIGH
            "primary_concern",
            "secondary_concerns",
            "developmental_status",
            "autism_screening_status",
            "behavioral_screening_status",
            "nutrition_status",
            "environment_status",
            "referral_needed",        # Yes/No
            "referral_urgency",       # Routine, Soon, Urgent
            "intervention_priority",  # LOW, MODERATE, HIGH, URGENT
            "recommended_actions",
            "follow_up_date",
            "assessment_date",
            "assessor_remarks"
        ]
        ws.append(headers)
        
        baseline = data.get("baseline_risk", {})
        
        # Determine referral urgency based on overall risk
        risk_category = baseline.get("overall_risk_category", "LOW")
        referral_urgency_map = {
            "LOW": "Routine",
            "MEDIUM": "Routine",
            "MEDIUM-HIGH": "Soon",
            "HIGH": "Urgent"
        }
        
        # Determine developmental status
        risk = data.get("risk", {})
        num_delays = risk.get("num_delays", 0)
        if num_delays == 0:
            dev_status = "Normal"
        elif num_delays == 1:
            dev_status = "Mild Delay"
        elif num_delays == 2:
            dev_status = "Moderate Delay"
        else:
            dev_status = "Severe Delay"
        
        # Get neuro data for autism and behavior status
        neuro = data.get("neuro_behavioral", {})
        autism_status = f"{neuro.get('autism_risk', 'Low')} Risk"
        behavior_status = f"{neuro.get('behavior_risk', 'Low')} Risk"
        
        # Get nutrition status
        nutrition = data.get("nutrition", {})
        nutrition_status = nutrition.get("nutrition_risk", "Low")
        
        # Get environment status
        env = data.get("environment", {})
        env_score = env.get("parent_child_interaction_score", 0) + env.get("home_stimulation_score", 0)
        env_status = "Adequate" if env_score >= 15 else "Needs Improvement"
        
        row = [
            data.get("child_id", ""),
            data.get("child_unique_id", ""),
            risk_category,
            baseline.get("primary_concern", ""),
            baseline.get("secondary_concerns", ""),
            dev_status,
            autism_status,
            behavior_status,
            nutrition_status,
            env_status,
            "Yes" if baseline.get("referral_needed") else "No",
            referral_urgency_map.get(risk_category, "Routine"),
            baseline.get("intervention_priority", "LOW"),
            "",  # recommended_actions
            "",  # follow_up_date
            data.get("assessment_date", datetime.now().strftime("%Y-%m-%d")),
            ""   # assessor_remarks
        ]
        ws.append(row)
        ExcelService._format_sheet(ws)
    
    @staticmethod
    def _format_sheet(ws):
        """Apply formatting to sheet"""
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data styling
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply to header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Apply to data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set adjusted width with max limit
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 12)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    @staticmethod
    def export_bulk_assessment(children_data: List[Dict[str, Any]], output_path: str) -> str:
        """
        Export multiple children's assessments to a single Excel file
        with summary sheet and individual detail sheets
        """
        wb = Workbook()
        
        # Create Summary sheet
        summary_ws = wb.active
        summary_ws.title = "Summary"
        ExcelService._write_summary_sheet(summary_ws, children_data)
        
        # Create individual sheets for each child (limited to first 10 for file size)
        for i, child_data in enumerate(children_data[:10]):
            sheet_name = f"Child_{child_data.get('child_unique_id', f'{i+1}')[:20]}"
            # Ensure sheet name is valid (max 31 chars, no special chars)
            sheet_name = sheet_name[:31].replace('/', '_').replace('\\', '_')
            
            child_ws = wb.create_sheet(sheet_name)
            
            # Write key data for this child
            headers = [
                "Metric", "Value"
            ]
            child_ws.append(headers)
            
            # Basic info
            child_ws.append(["Child ID", child_data.get("child_unique_id", "")])
            child_ws.append(["Age (months)", child_data.get("child_age_months", "")])
            child_ws.append(["Gender", child_data.get("gender", "")])
            child_ws.append(["Assessment Date", child_data.get("assessment_date", "")])
            child_ws.append([])
            
            # Developmental scores
            dev = child_data.get("developmental", {})
            child_ws.append(["GM DQ", dev.get("gm_dq", "")])
            child_ws.append(["FM DQ", dev.get("fm_dq", "")])
            child_ws.append(["LC DQ", dev.get("lc_dq", "")])
            child_ws.append(["COG DQ", dev.get("cog_dq", "")])
            child_ws.append(["SE DQ", dev.get("se_dq", "")])
            child_ws.append(["Composite DQ", dev.get("composite_dq", "")])
            child_ws.append([])
            
            # Risk
            baseline = child_data.get("baseline_risk", {})
            child_ws.append(["Overall Risk", baseline.get("overall_risk_category", "")])
            child_ws.append(["Referral Needed", "Yes" if baseline.get("referral_needed") else "No"])
            child_ws.append(["Intervention Priority", baseline.get("intervention_priority", "")])
            
            ExcelService._format_sheet(child_ws)
        
        wb.save(output_path)
        return output_path
    
    @staticmethod
    def _write_summary_sheet(ws, children_data: List[Dict[str, Any]]):
        """Write summary sheet with all children"""
        headers = [
            "S.No", "Child ID", "Age (months)", "Gender",
            "GM DQ", "FM DQ", "LC DQ", "COG DQ", "SE DQ", "Composite DQ",
            "Num Delays", "Autism Risk", "Behavior Risk", "Nutrition Risk",
            "Overall Risk", "Referral Needed", "Intervention Priority"
        ]
        ws.append(headers)
        
        for i, child in enumerate(children_data, 1):
            dev = child.get("developmental", {})
            risk = child.get("risk", {})
            neuro = child.get("neuro_behavioral", {})
            nutrition = child.get("nutrition", {})
            baseline = child.get("baseline_risk", {})
            
            row = [
                i,
                child.get("child_unique_id", ""),
                child.get("child_age_months", ""),
                child.get("gender", ""),
                dev.get("gm_dq", ""),
                dev.get("fm_dq", ""),
                dev.get("lc_dq", ""),
                dev.get("cog_dq", ""),
                dev.get("se_dq", ""),
                dev.get("composite_dq", ""),
                risk.get("num_delays", 0),
                neuro.get("autism_risk", ""),
                neuro.get("behavior_risk", ""),
                nutrition.get("nutrition_risk", ""),
                baseline.get("overall_risk_category", ""),
                "Yes" if baseline.get("referral_needed") else "No",
                baseline.get("intervention_priority", "")
            ]
            ws.append(row)
        
        ExcelService._format_sheet(ws)
